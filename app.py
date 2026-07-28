from flask import Flask, render_template, request, send_file, jsonify, Response, stream_with_context
import pandas as pd
import io
import os
import json
from werkzeug.utils import secure_filename
from datetime import datetime, date

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
LOG_FOLDER = 'sql_log'

if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_sql_template(template):
    now = datetime.now()
    filename = now.strftime('%Y%m%d%H%M%S%f')[:-3] + '.txt'
    filepath = os.path.join(LOG_FOLDER, filename)
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(template)
        return True
    except Exception as e:
        print(f'保存SQL模板失败: {e}')
        return False

def convert_value(value):
    """将pandas读取的值转为可JSON序列化的格式，同时保留长数字精度"""
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float):
        return str(value)
    if isinstance(value, (int)):
        return str(value)
    return str(value)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': '只支持Excel文件 (.xlsx, .xls)'}), 400
    
    def generate():
        try:
            import openpyxl
            file.stream.seek(0)
            wb = openpyxl.load_workbook(file.stream, read_only=True, data_only=True)
            ws = wb.active
            total = max((ws.max_row or 1) - 1, 0)
            yield json.dumps({'type': 'progress', 'loaded': 0, 'total': total}) + '\n'
            
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                yield json.dumps({'type': 'error', 'error': 'Excel文件为空'}) + '\n'
                return
            
            columns = [str(c) if c is not None else '' for c in header]
            columns = [c if c else f'column{i+1}' for i, c in enumerate(columns)]
            
            data = []
            loaded = 0
            for row in rows:
                row_data = {}
                for idx, col in enumerate(columns):
                    val = row[idx] if idx < len(row) else None
                    if val is None or (isinstance(val, str) and val.strip() == ''):
                        row_data[col] = None
                    elif isinstance(val, (datetime, date)):
                        row_data[col] = val.strftime('%Y-%m-%d')
                    else:
                        row_data[col] = str(val)
                data.append(row_data)
                loaded += 1
                yield json.dumps({'type': 'progress', 'loaded': loaded, 'total': total}) + '\n'
            
            wb.close()
            yield json.dumps({
                'type': 'done',
                'columns': columns,
                'data': data,
                'row_count': len(data),
                'col_count': len(columns)
            }) + '\n'
        except Exception as e:
            import traceback
            error_msg = f'解析Excel文件失败: {str(e)}\n{traceback.format_exc()}'
            print(error_msg)
            yield json.dumps({'type': 'error', 'error': f'解析Excel文件失败: {str(e)}'}) + '\n'
    
    return Response(stream_with_context(generate()), mimetype='application/x-ndjson')

@app.route('/generate_sql', methods=['POST'])
def generate_sql():
    try:
        data = request.get_json()
        template = data.get('template', '')
        excel_data = data.get('data', [])
        columns = data.get('columns', [])
        
        if not template:
            return jsonify({'error': '请输入SQL模板'}), 400
        
        if not excel_data:
            return jsonify({'error': '请先上传Excel文件'}), 400
        
        save_sql_template(template)
        
        sql_statements = []
        for row in excel_data:
            sql = template
            
            placeholders = []
            for idx, col in enumerate(columns):
                placeholder = f'$column{idx + 1}$'
                value = row.get(col)
                
                if value is None:
                    sql_value = 'NULL'
                else:
                    sql_value = str(value)
                
                placeholders.append((placeholder, sql_value))
            
            placeholders.sort(key=lambda x: len(x[0]), reverse=True)
            
            for placeholder, sql_value in placeholders:
                sql = sql.replace(placeholder, sql_value)
            
            sql_statements.append(sql)
        
        output = '\n'.join(sql_statements)
        preview_output = '\n'.join(sql_statements[:50])
        
        return jsonify({
            'sql': output,
            'preview_sql': preview_output,
            'preview_count': min(50, len(sql_statements)),
            'count': len(sql_statements)
        })
    except Exception as e:
        import traceback
        error_msg = f'生成SQL失败: {str(e)}\n{traceback.format_exc()}'
        print(error_msg)
        return jsonify({'error': f'生成SQL失败: {str(e)}'}), 400

@app.route('/download', methods=['POST'])
def download_sql():
    try:
        data = request.get_json()
        sql_content = data.get('sql', '')
        
        if not sql_content:
            return jsonify({'error': '没有SQL内容可下载'}), 400
        
        output = io.BytesIO()
        output.write(sql_content.encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='text/plain',
            as_attachment=True,
            download_name='generated_sql.txt'
        )
    except Exception as e:
        import traceback
        error_msg = f'下载失败: {str(e)}\n{traceback.format_exc()}'
        print(error_msg)
        return jsonify({'error': f'下载失败: {str(e)}'}), 400

if __name__ == '__main__':
    import sys
    _stderr_write = sys.stderr.write
    sys.stderr.write = lambda s: _stderr_write(s) if 'development server' not in s else None
    app.run(debug=False, host='0.0.0.0', port=5000)
